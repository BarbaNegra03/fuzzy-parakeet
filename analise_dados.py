import openpyxl

# Carregar a planilha de fechamento
planilha_fechamento = openpyxl.load_workbook(r'C:\Users\Gabriel\Desktop\consulta pagamentos\planilha_fechamento.xlsx')
pagina_fechamento = planilha_fechamento.active

# Contadores
em_dia = 0
pendente = 0

# Imprimir dados da planilha
print("Dados da planilha:")
for linha in pagina_fechamento.iter_rows(min_row=2, values_only=True):
    print(linha)

# Analisar os dados da planilha
for linha in pagina_fechamento.iter_rows(min_row=2, values_only=True):
    status = linha[4]
    print(f"Status: {status}")
    if status == 'em dia':
        em_dia += 1
    elif status == 'pendente':
        pendente += 1

# Gerar relatório simples
print(f"Clientes em dia: {em_dia}")
print(f"Clientes pendentes: {pendente}")
