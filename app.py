'''
1 - Entrar na planilha e extrair o cpf do cliente 
2 - Entro no site https://consultcpf-devaprender.netlify.app/ do cliente
3 - Verificar se está "Em dia", pegar a data do pagamento e o método de pagamento
4 - Se estiver "em dia", pegar a data do pagamento e o método de pagamento
5 - Caso contrário (se estiver atrasado), colocar o status como pedente
6 - Inserir essas informações(nome valor, cpf, vencimento, status e caso esteja em dia, data pagamento) em uma nova planilha, método pagamento(cartao ou boleto)) em uma nova planilha
7 - Repetir até chegar no último cliente
'''
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep

# Inicializar o WebDriver
service = ChromeService(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

print("WebDriver inicializado")

# Carregar a planilha de dados
planilha_clientes = openpyxl.load_workbook(r'C:\Users\Gabriel\Desktop\consulta pagamentos\dados_clientes_novo.xlsx')
print("Planilha carregada")

# Acessar a página (aba) desejada, por exemplo, a aba chamada "Sheet1"
pagina_clientes = planilha_clientes['Sheet1']

# Lista para armazenar os resultados
resultados = []

# 1 - Entrar na planilha e extrair os dados do cliente
for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha
    print(f"Processando: {nome}, {cpf}")
    print(valor)
    print(cpf)
    print(vencimento)

    # Abrir o site para cada CPF
    print("Abrindo site de consulta")
    driver.get('https://consultcpf-devaprender.netlify.app/')
    sleep(5)

    print("Site aberto")

    # Preencher o campo de pesquisa
    print("Preenchendo campo de pesquisa")
    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfinput']")
    sleep(1)
    campo_pesquisa.send_keys(cpf)
    sleep(1)

    # Clicar no botão de pesquisa
    print("Clicando no botão de pesquisa")
    botao_pesquisar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-block mt-3']")
    sleep(1)
    botao_pesquisar.click()
    sleep(4)

    # Verificar o status
    print("Verificando status")
    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']").text
    print(f"Status: {status}")

    if status == 'em dia':
        try:
            # 4 - Se estiver "em dia", pegar a data do pagamento e o método de pagamento
            data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']").text
            metodo_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']").text
        except:
            data_pagamento = "Não encontrado"
            metodo_pagamento = "Não encontrado"
        resultados.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento, metodo_pagamento])
    else:
        # 5 - Caso contrário (se estiver atrasado), colocar o status como pendente
        resultados.append([nome, valor, cpf, vencimento, 'pendente', '', ''])

print("Processamento concluído")

# Fechar o WebDriver no final do loop
driver.quit()
print("WebDriver fechado")

# Salvar os resultados na planilha de fechamento
planilha_fechamento = openpyxl.Workbook()
pagina_fechamento = planilha_fechamento.active
pagina_fechamento.append(['Nome', 'Valor', 'CPF', 'Vencimento', 'Status', 'Data Pagamento', 'Método Pagamento'])

for resultado in resultados:
    pagina_fechamento.append(resultado)

planilha_fechamento.save(r'C:\Users\Gabriel\Desktop\consulta pagamentos\planilha_fechamento.xlsx')
print("Resultados salvos na planilha de fechamento")

